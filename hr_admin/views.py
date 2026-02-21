from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.models import User
from recruitment.models import (
    Vacancy, Application, Document, Interview, InterviewScore,
    Notification, BulkMessage, CATEGORIES, QUALIFICATION_LEVELS, APPLICATION_STATUS
)
from accounts.models import PROVINCES, ROLE_HR_ADMIN
from .forms import VacancyForm, ApplicationFilterForm, BulkMessageForm, InterviewScheduleForm
import json
from datetime import date


def hr_required(func):
    """Decorator to restrict access to HR admins only."""
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            if request.user.profile.role not in [ROLE_HR_ADMIN, 'hr_admin'] and not request.user.is_superuser:
                messages.error(request, 'Access denied. HR Administrator role required.')
                return redirect('recruitment:job_list')
        except Exception:
            messages.error(request, 'Access denied.')
            return redirect('recruitment:job_list')
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


@hr_required
def dashboard(request):
    vacancies = Vacancy.objects.all()
    applications = Application.objects.all()
    today = date.today()

    stats = {
        'total_vacancies': vacancies.count(),
        'open_vacancies': vacancies.filter(status='open').count(),
        'total_applications': applications.count(),
        'new_today': applications.filter(submitted_at__date=today).count(),
        'under_review': applications.filter(status='under_review').count(),
        'shortlisted': applications.filter(status='shortlisted').count(),
        'selected': applications.filter(status='selected').count(),
        'interviews_today': Interview.objects.filter(scheduled_date__date=today, status='scheduled').count(),
    }

    recent_applications = applications.select_related('vacancy', 'applicant').order_by('-submitted_at')[:10]
    open_vacancies = vacancies.filter(status='open').annotate(app_count=Count('applications'))

    # Province distribution for chart
    province_data = applications.values('province').annotate(count=Count('id')).order_by('-count')[:10]
    province_labels = [p['province'] for p in province_data]
    province_counts = [p['count'] for p in province_data]

    # Status distribution
    status_data = applications.values('status').annotate(count=Count('id'))
    status_labels = [s['status'].replace('_', ' ').title() for s in status_data]
    status_counts = [s['count'] for s in status_data]

    return render(request, 'hr_admin/dashboard.html', {
        'stats': stats,
        'recent_applications': recent_applications,
        'open_vacancies': open_vacancies,
        'province_labels': json.dumps(province_labels),
        'province_counts': json.dumps(province_counts),
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
    })


# ---------- Vacancy Management ----------

@hr_required
def vacancy_list(request):
    vacancies = Vacancy.objects.annotate(app_count=Count('applications')).order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        vacancies = vacancies.filter(status=status_filter)
    return render(request, 'hr_admin/vacancy_list.html', {
        'vacancies': vacancies,
        'status_filter': status_filter,
    })


@hr_required
def vacancy_create(request):
    if request.method == 'POST':
        form = VacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save(commit=False)
            vacancy.created_by = request.user
            vacancy.save()
            messages.success(request, f'Vacancy "{vacancy.title}" created successfully.')
            return redirect('hr_admin:vacancy_list')
    else:
        form = VacancyForm()
    return render(request, 'hr_admin/vacancy_form.html', {'form': form, 'title': 'Create Vacancy'})


@hr_required
def vacancy_edit(request, pk):
    vacancy = get_object_or_404(Vacancy, pk=pk)
    if request.method == 'POST':
        form = VacancyForm(request.POST, instance=vacancy)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vacancy updated successfully.')
            return redirect('hr_admin:vacancy_list')
    else:
        form = VacancyForm(instance=vacancy)
    return render(request, 'hr_admin/vacancy_form.html', {'form': form, 'title': 'Edit Vacancy', 'vacancy': vacancy})


@hr_required
def vacancy_toggle_status(request, pk):
    vacancy = get_object_or_404(Vacancy, pk=pk)
    if vacancy.status == 'open':
        vacancy.status = 'closed'
        messages.info(request, f'Vacancy "{vacancy.title}" has been closed.')
    elif vacancy.status in ['draft', 'closed']:
        vacancy.status = 'open'
        messages.success(request, f'Vacancy "{vacancy.title}" is now open.')
    vacancy.save()
    return redirect('hr_admin:vacancy_list')


# ---------- Application Management ----------

@hr_required
def application_list(request):
    applications = Application.objects.select_related('vacancy', 'applicant').order_by('-submitted_at')

    vacancy_filter = request.GET.get('vacancy', '')
    status_filter = request.GET.get('status', '')
    province_filter = request.GET.get('province', '')
    search = request.GET.get('search', '')

    if vacancy_filter:
        applications = applications.filter(vacancy_id=vacancy_filter)
    if status_filter:
        applications = applications.filter(status=status_filter)
    if province_filter:
        applications = applications.filter(province=province_filter)
    if search:
        applications = applications.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search) |
            Q(email__icontains=search) | Q(vacancy__title__icontains=search)
        )

    vacancies = Vacancy.objects.all()
    return render(request, 'hr_admin/application_list.html', {
        'applications': applications,
        'vacancies': vacancies,
        'provinces': PROVINCES,
        'statuses': APPLICATION_STATUS,
        'vacancy_filter': vacancy_filter,
        'status_filter': status_filter,
        'province_filter': province_filter,
        'search': search,
    })


@hr_required
def application_detail(request, pk):
    application = get_object_or_404(Application, pk=pk)
    documents = application.documents.all()
    interviews = application.interviews.select_related().prefetch_related('panel_members', 'panel_scores').all()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_status':
            new_status = request.POST.get('status')
            if new_status in dict(APPLICATION_STATUS):
                old_status = application.status
                application.status = new_status
                application.hr_notes = request.POST.get('hr_notes', application.hr_notes)
                application.save()

                # Notify applicant
                status_display = dict(APPLICATION_STATUS).get(new_status, new_status)
                Notification.objects.create(
                    user=application.applicant,
                    title=f'Application Status Update',
                    message=f'Your application for {application.vacancy.title} has been updated to: {status_display}.',
                    notification_type='status_update',
                )
                messages.success(request, f'Status updated to {status_display}.')

        elif action == 'rescore':
            score = application.compute_score()
            messages.success(request, f'Application rescored: {score}/100')

        elif action == 'verify_doc':
            doc_id = request.POST.get('doc_id')
            doc = get_object_or_404(Document, pk=doc_id, application=application)
            doc.verified = True
            doc.verified_by = request.user
            doc.verified_at = timezone.now()
            doc.save()
            messages.success(request, 'Document marked as verified.')

        return redirect('hr_admin:application_detail', pk=pk)

    return render(request, 'hr_admin/application_detail.html', {
        'application': application,
        'documents': documents,
        'interviews': interviews,
        'statuses': APPLICATION_STATUS,
    })


@hr_required
def run_auto_screening(request, vacancy_pk):
    vacancy = get_object_or_404(Vacancy, pk=vacancy_pk)
    applications = vacancy.applications.all()
    scored = 0
    for app in applications:
        app.compute_score()
        scored += 1
    messages.success(request, f'Auto-screening complete. Scored {scored} applications for "{vacancy.title}".')
    return redirect('hr_admin:application_list')


@hr_required
def shortlist_view(request, vacancy_pk):
    vacancy = get_object_or_404(Vacancy, pk=vacancy_pk)
    applications = vacancy.applications.filter(
        total_score__isnull=False
    ).order_by('-total_score')

    threshold = float(request.GET.get('threshold', 50))

    if request.method == 'POST' and request.POST.get('action') == 'shortlist':
        ids = request.POST.getlist('selected_ids')
        Application.objects.filter(id__in=ids, vacancy=vacancy).update(status='shortlisted')
        # Notify rejected
        Application.objects.filter(vacancy=vacancy).exclude(id__in=ids).exclude(
            status__in=['shortlisted', 'interview_scheduled', 'interviewed', 'selected', 'withdrawn']
        ).update(status='rejected')

        count = len(ids)
        # Notify shortlisted applicants
        for app in Application.objects.filter(id__in=ids):
            Notification.objects.create(
                user=app.applicant,
                title='Congratulations - You Have Been Shortlisted!',
                message=f'We are pleased to inform you that your application for {vacancy.title} '
                        f'has been shortlisted. Further details regarding the interview will be provided shortly.',
                notification_type='success',
            )

        messages.success(request, f'{count} applicants shortlisted and notified.')
        return redirect('hr_admin:shortlist', vacancy_pk=vacancy_pk)

    qualified = applications.filter(total_score__gte=threshold)
    below_threshold = applications.filter(total_score__lt=threshold)

    return render(request, 'hr_admin/shortlist.html', {
        'vacancy': vacancy,
        'qualified': qualified,
        'below_threshold': below_threshold,
        'threshold': threshold,
    })


@hr_required
def bulk_message(request):
    if request.method == 'POST':
        form = BulkMessageForm(request.POST)
        if form.is_valid():
            vacancy = form.cleaned_data.get('vacancy')
            status_filter = form.cleaned_data.get('recipient_status')
            subject = form.cleaned_data['subject']
            message_body = form.cleaned_data['message']

            applications = Application.objects.select_related('applicant')
            if vacancy:
                applications = applications.filter(vacancy=vacancy)
            if status_filter:
                applications = applications.filter(status=status_filter)

            count = 0
            for app in applications:
                Notification.objects.create(
                    user=app.applicant,
                    title=subject,
                    message=message_body,
                    notification_type='info',
                )
                count += 1

            BulkMessage.objects.create(
                vacancy=vacancy,
                subject=subject,
                message=message_body,
                recipient_status=status_filter or '',
                sent_by=request.user,
                recipient_count=count,
            )

            messages.success(request, f'Message sent to {count} applicants.')
            return redirect('hr_admin:bulk_message')
    else:
        form = BulkMessageForm()

    recent = BulkMessage.objects.order_by('-sent_at')[:10]
    return render(request, 'hr_admin/bulk_message.html', {'form': form, 'recent': recent})


@hr_required
def interview_schedule(request, application_pk):
    application = get_object_or_404(Application, pk=application_pk)
    if request.method == 'POST':
        form = InterviewScheduleForm(request.POST)
        if form.is_valid():
            interview = form.save(commit=False)
            interview.application = application
            interview.save()
            form.save_m2m()

            application.status = 'interview_scheduled'
            application.save()

            Notification.objects.create(
                user=application.applicant,
                title='Interview Scheduled',
                message=f'Your interview for {application.vacancy.title} has been scheduled for '
                        f'{interview.scheduled_date.strftime("%A, %d %B %Y at %I:%M %p")} '
                        f'at {interview.venue}. Please arrive 15 minutes early.',
                notification_type='interview',
            )

            messages.success(request, 'Interview scheduled and applicant notified.')
            return redirect('hr_admin:application_detail', pk=application_pk)
    else:
        form = InterviewScheduleForm()

    return render(request, 'hr_admin/interview_schedule.html', {
        'form': form,
        'application': application,
    })


@hr_required
def reports(request):
    applications = Application.objects.all()
    vacancy_filter = request.GET.get('vacancy', '')
    if vacancy_filter:
        applications = applications.filter(vacancy_id=vacancy_filter)

    by_province = applications.values('province').annotate(count=Count('id')).order_by('-count')
    by_gender = applications.values('gender').annotate(count=Count('id'))
    by_status = applications.values('status').annotate(count=Count('id'))
    by_qualification = applications.values('highest_qualification').annotate(count=Count('id'))
    avg_score = applications.aggregate(avg=Avg('total_score'))

    province_labels = json.dumps([p['province'] for p in by_province])
    province_counts = json.dumps([p['count'] for p in by_province])

    return render(request, 'hr_admin/reports.html', {
        'by_province': by_province,
        'by_gender': by_gender,
        'by_status': by_status,
        'by_qualification': by_qualification,
        'avg_score': avg_score['avg'],
        'vacancies': Vacancy.objects.all(),
        'vacancy_filter': vacancy_filter,
        'province_labels': province_labels,
        'province_counts': province_counts,
        'total_applications': applications.count(),
    })


@hr_required
def export_shortlist(request, vacancy_pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    vacancy = get_object_or_404(Vacancy, pk=vacancy_pk)
    applications = vacancy.applications.filter(
        status__in=['shortlisted', 'interview_scheduled', 'interviewed', 'selected']
    ).order_by('-total_score')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shortlist'

    header_fill = PatternFill(start_color='003087', end_color='003087', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['#', 'Full Name', 'Gender', 'Province', 'Qualification', 'Institution',
               'Grade', 'Experience (yrs)', 'Phone', 'Email', 'Score', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, app in enumerate(applications, 1):
        ws.append([
            i,
            app.full_name(),
            app.gender,
            app.province,
            app.get_highest_qualification_display(),
            app.institution,
            app.grade_result,
            app.years_experience,
            app.phone,
            app.email,
            app.total_score or 0,
            app.get_status_display(),
        ])

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="shortlist_{vacancy.reference_number}.xlsx"'
    wb.save(response)
    return response
